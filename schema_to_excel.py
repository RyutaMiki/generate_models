# schema_to_excel.py – YAML → Excel table-definition workbook generator (v3.2)
"""Generate a polished Excel workbook (one sheet per table) from a YAML schema.

Changes in **v3.2**
-------------------

Usage
-----
$ python schema_to_excel.py schema.yaml [tables.xlsx]
"""
from __future__ import annotations

import re
import sys
from pathlib import Path
from typing import List

import pandas as pd
import yaml
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ----------------------------------------------------------------------------
# Configuration
# ----------------------------------------------------------------------------

TYPE_LIST: List[str] = [
    "String", "Text", "Integer", "Float", "Boolean", "Date", "TIMESTAMP",
    "DECIMAL", "SmallInteger", "EnumType",
]

BOOL_COLUMNS = {"Nullable", "Primary Key", "Autoincrement", "Unique", "Index"}
BOOL_DV_FORMULA = '"TRUE,FALSE"'

HEADER_ORDER = [
    "No", "Column Name", "Type", "Args", "Nullable", "Primary Key",
    "Autoincrement", "Default", "On Update", "Server Default", "Unique",
    "Index", "Comment",
]

INVALID_CHRS = re.compile(r"[\\/?*\[\]:]")

# Styles
HDR_FILL = PatternFill("solid", fgColor="4F81BD")
HDR_FONT = Font(bold=True, color="FFFFFF")
THIN = Side(style="thin", color="A6A6A6")
BORDER = Border(top=THIN, left=THIN, right=THIN, bottom=THIN)
ALT_FILL = PatternFill("solid", fgColor="F2F2F2")
CENTER = Alignment(horizontal="center", vertical="center")

# ----------------------------------------------------------------------------
# Helpers
# ----------------------------------------------------------------------------

def _safe_sheet_name(name: str, used: set[str]) -> str:
    safe = INVALID_CHRS.sub("_", name)
    safe = re.sub(r"[^0-9A-Za-z_]+", "_", safe)[:31] or "Sheet"
    candidate = safe
    i = 1
    while candidate in used:
        suf = f"_{i}"
        candidate = safe[: 31 - len(suf)] + suf
        i += 1
    used.add(candidate)
    return candidate


def _build_df(table: dict) -> pd.DataFrame:
    rows = []
    for i, col in enumerate(table["columns"], 1):
        rows.append({
            "No": i,
            "Column Name": col["name"],
            "Type": col["type"],
            "Args": ", ".join(map(str, col.get("args", []))) if col.get("args") else "",
            "Nullable": col.get("nullable", True),
            "Primary Key": col.get("primary_key", False),
            "Autoincrement": col.get("autoincrement", False),
            "Default": col.get("default", ""),
            "On Update": col.get("onupdate", ""),
            "Server Default": col.get("server_default", ""),
            "Unique": col.get("unique", False),
            "Index": col.get("index", False),
            "Comment": col.get("comment", ""),
        })
    return pd.DataFrame(rows, columns=HEADER_ORDER)

# ----------------------------------------------------------------------------
# Core
# ----------------------------------------------------------------------------

def write_excel(schema: dict, out_path: Path) -> None:
    """Create Excel workbook from schema with styled sheets."""
    used_names: set[str] = set()
    sheet_meta: dict[str, dict] = {}

    # 1️⃣ pandas -> openpyxl for raw data
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        for model in schema["models"]:
            sheet_name = _safe_sheet_name(model["class_name"], used_names)
            _build_df(model).to_excel(writer, sheet_name=sheet_name, index=False)
            sheet_meta[sheet_name] = {
                "table_name": model["table_name"],
                "description": str(model.get("description", "")).strip(),
            }

    # 2️⃣ Styling, dropdowns, titles
    wb = load_workbook(out_path)
    for ws in wb.worksheets:
        meta = sheet_meta[ws.title]

        # Insert two rows at top for titles
        ws.insert_rows(1, amount=2)
        ws["A1"] = meta["table_name"]  # テーブル物理名
        ws["A2"] = meta["description"]  # テーブル和名 / 説明
        title_font = Font(bold=True, size=13)
        ws["A1"].font = title_font
        ws["A2"].font = Font(bold=False, size=12)

        ws.row_dimensions[1].height = 24
        ws.row_dimensions[2].height = 20

        max_col_len = {}

        # Header row is now row 3
        header_row_idx = 3
        ws.row_dimensions[header_row_idx].height = 22
        for cell in ws[header_row_idx]:
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.alignment = CENTER
            cell.border = BORDER
            max_col_len[cell.column] = len(str(cell.value))

        # Body styling starting row 4
        start_body = header_row_idx + 1
        for row in ws.iter_rows(min_row=start_body):
            r_idx = row[0].row
            ws.row_dimensions[r_idx].height = 18
            if (r_idx - start_body) % 2 == 0:
                for cell in row:
                    cell.fill = ALT_FILL
            for cell in row:
                cell.border = BORDER
                max_col_len[cell.column] = max(
                    max_col_len.get(cell.column, 0),
                    len(str(cell.value)) if cell.value is not None else 0,
                )

        # Column width calculation
        for col_idx, length in max_col_len.items():
            header_val = ws.cell(row=header_row_idx, column=col_idx).value
            if header_val == "Comment":
                width = min(max(length + 15, 40), 100)
            else:
                width = min(max(length + 6, 14), 45)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Dropdowns
        header_map = {ws.cell(row=header_row_idx, column=col).value: get_column_letter(col)
                      for col in range(1, ws.max_column + 1)}
        type_col = header_map.get("Type")
        if type_col:
            dv = DataValidation(type="list", formula1='"' + ",".join(TYPE_LIST) + '"', allow_blank=False)
            ws.add_data_validation(dv)
            dv.add(f"{type_col}{start_body}:{type_col}1048576")
        for field in BOOL_COLUMNS:
            col_letter = header_map.get(field)
            if col_letter:
                dv_bool = DataValidation(type="list", formula1=BOOL_DV_FORMULA, allow_blank=False)
                ws.add_data_validation(dv_bool)
                dv_bool.add(f"{col_letter}{start_body}:{col_letter}1048576")

        # Freeze: keep titles & header visible
        ws.freeze_panes = f"A{start_body}"

    wb.save(out_path)

# ----------------------------------------------------------------------------
# CLI
# ----------------------------------------------------------------------------

if __name__ == "__main__":
    try:
        if len(sys.argv) not in (2, 3):
            raise SystemExit("Usage: python schema_to_excel.py <schema.yaml> [output.xlsx]")
        schema_path = Path(sys.argv[1])
        out_path = Path(sys.argv[2]) if len(sys.argv) == 3 else schema_path.with_suffix(".xlsx")
        if not schema_path.exists():
            raise SystemExit(f"Schema file not found: {schema_path}")
        with schema_path.open(encoding="utf-8") as f:
            schema = yaml.safe_load(f)
        write_excel(schema, out_path)
        print(f"Excel workbook written → {out_path.resolve()}")
    except Exception as err:  # noqa: BLE001
        print("Error:", err, file=sys.stderr)
        sys.exit(1)
